'''
Last modified on : 19-oct-2020
Author: Devi lalitamba paidi & Dasari Vamsi
'''

import sys
import os
from tkinter import *
import tkinter as tk
import tkinter.ttk as ttk
import time
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox  
import pandas as pd
import operator
import math
import xlsxwriter
import xlwings as xw
import datetime
import time
import PIL.Image
import PIL.ImageTk
import xlrd as xl
from pandas import ExcelWriter
from pandas import ExcelFile
from pandas import DataFrame
import win32com.client as client
import pathlib
import pandas as pd
from openpyxl import load_workbook
import shutil
from zipfile import ZipFile
import os
from datetime import date
from datetime import timedelta
from os.path import basename
#currenttime = time.strftime("%d-%B-%Y %I:%M:%S:%p")
currenttime = time.strftime("%B-%d-%Y_%I_%M_%S_%p")
root = tk.Tk() 
# Create a progressbar widget
#progress_bar = ttk.Progressbar(root, orient="horizontal",length=420,mode="determinate", maximum=100, value=0) 
# And a label for it
#label_1 = tk.Label(root, text="Progress Bar")
# Use the grid manager
#label_1.grid(row=80, column=0)
#progress_bar.grid(row=80, column=1) 
#progress_bar.place(x=80,y=350)
# Necessary, as the root object needs to draw the progressbar widget
# Otherwise, it will not be visible on the screen
#root.update() 
#progress_bar['value'] = 0
root.update()

root.title('IQVIA Daily Report Generation')
root.geometry("600x500") #You want the size of the app to be 500x500
root.resizable(0, 0) #Don't allow resizing in the x or y direction
#label = tk.Label(root,text="RU Billing Report Generation Tool V1.0",justify=tk.LEFT,fg = "blue",font = "Helvetica 16 bold italic").pack()
label = Label(text="IQVIA Daily Report Generation",fg = "blue",font = "Helvetica 15 bold italic")
label.place(x=180,y=10)

im = PIL.Image.open("C:\\Users\\vamsi.dasari\\Desktop\\test1\\dwp.JPG")
photo = PIL.ImageTk.PhotoImage(im)
label = Label(root, image=photo)
#label.image = photo # keep a reference!
label.place(x=10,y=10)

text = Label(text="Browse Input Files to Process",font = "Helvetica 14 bold italic")
text.place(x=150,y=60)
cmplt_order_track = Label(text="completed_orders_file",fg = "Black",font = "Helvetica 15 bold italic")
cmplt_order_track.place(x=11,y=130)
#------- complete_Orders_Tracking File ----------#
def browsefunc():
    global cmplt_order_tracking_file
    cmplt_order_tracking_file = filedialog.askopenfilename(initialdir = "*",title = "Choose complete_Orders_Tracking File",filetypes = (("xlsx files","*.xlsx"),("all files","*.*")))
    if cmplt_order_tracking_file == "":
        messagebox.showwarning("warning","Please select complete_Orders_Tracking File")      
        #raise SystemExit 
    else:
        cmplt_report = os.path.basename(cmplt_order_tracking_file) 
        input1['text'] = cmplt_report

input1 = Label(text="____________________",fg ="blue",font = "Helvetica")
input1.place(x=240,y=130)
browsebutton1 = Button(text="Browse",fg ="blue",font = "Helvetica",command=browsefunc)
browsebutton1.place(x = 480,y = 130)
#-----------------------------------------#
termination = Label(text="Termination_orders_file",fg = "Black",font = "Helvetica 15 bold italic")
termination.place(x=10,y=170)
#-------Termination_orders_return_file--------------------#
def browsefunc1():
    global termination_file
    termination_file = filedialog.askopenfilename(initialdir = "*",title = "Choose Termination_orders_return_file",filetypes = (("xlsx files","*.xlsx"),("all files","*.*")))
    if termination_file == "":
        messagebox.showwarning("warning","Please select Termination_orders_return File")      
        #raise SystemExit
    else:
        termination_report = os.path.basename(termination_file) 
        input2['text'] = termination_report

input2 = Label(text="____________________",fg ="blue",font = "Helvetica")
input2.place(x=240,y=170)
browsebutton2 = Button(text="Browse",fg="blue",font = "Helvetica",command=browsefunc1)
browsebutton2.place(x = 480,y = 170)
#-----------------------------------------#
shipping = Label(text="Shipping_items_file",fg = "Black",font = "Helvetica 15 bold italic")
shipping.place(x=10,y=210)
#-------Shipping_Items Report--------------------#
def browsefunc2():
    global shipped_items
    shipped_items = filedialog.askopenfilename(initialdir = "*",title = "Choose Shipping_Items xlsx File",filetypes = (("xlsx files","*.xlsx"),("all files","*.*")))
    if shipped_items == "":
        messagebox.showwarning("warning","Please select Shipping_Items Report File")      
        #raise SystemExit
    else:
        shipped_items_name = os.path.basename(shipped_items) 
        input3['text'] = shipped_items_name

input3 = Label(text="____________________",fg ="blue",font = "Helvetica")
input3.place(x=240,y=210)
browsebutton3 = Button(text="Browse",fg="blue",font = "Helvetica",command=browsefunc2)
browsebutton3.place(x = 480,y = 210)
#-----------------------------------------#
inventory = Label(text="Inventory_file",fg = "Black",font = "Helvetica 15 bold italic")
inventory.place(x=10,y=250)
#-------inventoty_file--------------------#
def browsefunc3():
    global inventoty_file
    inventoty_file = filedialog.askopenfilename(initialdir = "*",title = "Select inventoty_file xlsx File",filetypes = (("xlsx files","*.xlsx"),("all files","*.*")))
    if inventoty_file == "":
        messagebox.showwarning("warning","Please select inventoty  File")      
        #raise SystemExit
    else:
        inventoty_file_name = os.path.basename(inventoty_file) 
        input4['text'] = inventoty_file_name = os.path.basename(inventoty_file) 

input4 = Label(text="____________________",fg ="blue",font = "Helvetica")
input4.place(x=240,y=250)
browsebutton4 = Button(text="Browse",fg="blue",font = "Helvetica",command=browsefunc3)
browsebutton4.place(x = 480,y = 250)
#-----------------------------------------#
Enter_Date = Label(text="Enter_Date",fg = "Black",font = "Helvetica 15 bold italic")
Enter_Date.place(x=10,y=290)
date_entry = tk.Entry (root)
date_entry.place(x=240,y=290)
Enter_Da = Label(text="YYYY-MM-DD",fg = "blue",font = "Helvetica 12 bold italic")
Enter_Da.place(x=420,y=290)

text = Label(text="Once files are selected, Click On Generate Report Button",font = "Helvetica 12 bold italic")
text.place(x=75,y=330)
def GenerateReport():   
    MsgBox = messagebox.askquestion('Start Application','Save and close all Excel Files before start the tool',icon = 'warning')
    if MsgBox == 'no':
        raise SystemExit
    
    if cmplt_order_tracking_file != "" and termination_file != "" and shipped_items != "" and inventoty_file != "":
        ####################Completed tracing orders updation#########################################
        #reading the data from excel_sheet1
        yesterday = date_entry.get()
        print(yesterday)
        DF1=pd.read_excel(cmplt_order_tracking_file,sheet_name='Ship')
        ship = DF1[['Line Item Ship Date','Order #','Sponsor','Study','Site Country','Site Number','Site Name','Quantity Shipped','Product Asset #','Order Sub-Category','Product Name','Shipment Waybill #','Courier']]

        with pd.ExcelWriter(('C:\\Users\\vamsi.dasari\\Desktop\\test1\\example.xlsx' )) as writer:
            ship.to_excel(writer, sheet_name='sheet1',index=False)

        ship = pd.read_excel('C:\\Users\\vamsi.dasari\\Desktop\\test1\\example.xlsx',sheet_name= 'sheet1')
        ship['Line Item Ship Date']=pd.to_datetime(ship['Line Item Ship Date'])
        print("....")
        shipd = ship.sort_values(by = "Line Item Ship Date")

        with pd.ExcelWriter(('C:\\Users\\vamsi.dasari\\Desktop\\test1\\example.xlsx' )) as writer:
            shipd.to_excel(writer, sheet_name='sheet1',index=False)
            
        group = shipd.groupby(by = shipd['Line Item Ship Date'].dt.date)
        #print(group)
        print("....")
        for dat,sponsor in group:
            #print(dat)
            if( str(dat) == str(yesterday)):
                #print(sponsor)
                latest1 = pd.DataFrame(sponsor)
    
        #print(latest)
    
        try:
            with pd.ExcelWriter(('C:\\Users\\vamsi.dasari\\Desktop\\test1\\latestrecords.xlsx' )) as writer:
                latest1.to_excel(writer, sheet_name='sheet1',index=False)
            rec = pd.read_excel('C:\\Users\\vamsi.dasari\\Desktop\\test1\\latestrecords.xlsx',sheet_name= 'sheet1')
        except:
            
            print("NO COMPLETE ORDER TRACKING ORDERS")
            try:
                print("Creating Sponser report with RMA Termination Tab")
                print("....")
                DF=pd.read_excel(termination_file ,sheet_name="Order Contact")

                Rma_ship= DF[["Line Item Ship Date","Order #","Order Sub-Category","Sponsor","Study","Site Number","Product Name"]]

                Rma_ship['Line Item Ship Date']=pd.to_datetime(Rma_ship['Line Item Ship Date'])

                RMA = Rma_ship.sort_values(by = "Line Item Ship Date")

                group = RMA.groupby(by = RMA['Line Item Ship Date'].dt.date)
                print("....")
                for dat,sponsor in group:
                   #print(dat)
                    if( str(dat) == str(yesterday)):
                        #print(sponsor)
                        latest2 = pd.DataFrame(sponsor)
    
                #print(latest)
                with pd.ExcelWriter(('C:\\Users\\vamsi.dasari\\Desktop\\test1\\RMA_latestrecords.xlsx' )) as writer:
                    latest2.to_excel(writer, sheet_name='sheet1',index=False)
                rec = pd.read_excel('C:\\Users\\vamsi.dasari\\Desktop\\test1\\RMA_latestrecords.xlsx',sheet_name= 'sheet1')
            except:
                print("NO RMA TERMINATION ORDERS")
                try:
                    print("Creating Sponser report with Shipped Sent Items Tab")
                    print("....")
                    DF=pd.read_excel(shipped_items ,sheet_name="Orders Shipped")

                    shipped= DF[["Shipment Date","Order #","Order Sub-Category","Study","Shipment Waybill #","Site Name","Site Number","Product Name","Sponsor"]]

                    shipped['Shipment Date']=pd.to_datetime(shipped['Shipment Date'])

                    SHIP = shipped.sort_values(by = "Shipment Date")

                    group1 = SHIP.groupby(by = SHIP['Shipment Date'].dt.date)

                    for dat,sponsor in group1:
                        #print(dat)
                        if( str(dat) == str(yesterday)):
                            #print(sponsor)
                            latest = pd.DataFrame(sponsor)
    
                    #print(latest)
                    with pd.ExcelWriter(('C:\\Users\\vamsi.dasari\\Desktop\\test1\\sent_latestrecords.xlsx' )) as writer:
                        latest.to_excel(writer, sheet_name='sheet1',index=False)
                    rec1 = pd.read_excel('C:\\Users\\vamsi.dasari\\Desktop\\test1\\sent_latestrecords.xlsx',sheet_name= 'sheet1')
                except:
                    msg="Latest Records are not found on "+str(yesterday)+",please close the widget"
                    MsgBox1 = messagebox.askquestion('Exit Application',msg,icon = 'warning')
                else:
                    index_names1 = rec1[ rec1['Product Name'] == 'IQVIA Simon TMO SIM Card' ].index
                    rec1.drop(index_names1, inplace = True)
                    with pd.ExcelWriter(('C:\\Users\\vamsi.dasari\\Desktop\\test1\\sent_latestrecords.xlsx' )) as writer:
                        rec1.to_excel(writer, sheet_name='latest',index=False)
                    sponsor_ship = pd.read_excel('C:\\Users\\vamsi.dasari\\Desktop\\test1\\sent_latestrecords.xlsx',sheet_name= 'latest')

                    sponsor_data_ship = sponsor_ship.groupby(by = sponsor_ship['Sponsor'])

                    for group_n,group_d in sponsor_data_ship:
                        #print(group_name)
                        #print(group_data)
                        study = group_d.groupby(by = group_d["Study"])
                        for group_name,group_data in study:
                            path = 'C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy'
                            entries = os.listdir(path)
                            #print(entries)
                            stri = group_name
                            name = stri.split('/')
                            group = (str(group_n)+ str(name[0]) +'.xlsx')
                            #print(group)
                            if(group in entries):
                                check = pd.ExcelFile('C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy\\'+str(group))
                                sheetnames = check.sheet_names
                                if('Completed Send Items Orders' in sheetnames):
                                    writer = pd.ExcelWriter(("C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy\\"+ str(group)),engine = 'openpyxl')
                                    writer.book = load_workbook("C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy\\"+ str(group))
                                    writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
                                    reader = pd.read_excel(("C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy\\"+ str(group)),sheet_name = 'Completed Send Items Orders')
                                    read = reader.index
                                    group_data.to_excel(writer,index=False,header=False,sheet_name = 'Completed Send Items Orders',startrow=len(read)+1)
                                    writer.close()
                            else:
                                src = 'C:\\Users\\vamsi.dasari\\Desktop\\test1\\summary_basesheet.xlsx'
                                dst = ('C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy\\' + str(group))
                                shutil.copyfile(src, dst)
                                with pd.ExcelWriter(('C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy\\'+ str(group)),engine='openpyxl', mode='a') as writer:
                                    group_data.to_excel(writer, sheet_name='Completed Send Items Orders',index=False)
                    print("....")
                    print("Created Sponser report with Shipped Sent Items Tab")
                
            else:
                index_names = rec[ rec['Product Name'] == 'IQVIA Simon TMO SIM Card' ].index
                rec.drop(index_names, inplace = True)
                with pd.ExcelWriter(('C:\\Users\\vamsi.dasari\\Desktop\\test1\\RMA_latestrecords.xlsx' )) as writer:
                    rec.to_excel(writer, sheet_name='latest',index=False)
                sponsor1 = pd.read_excel('C:\\Users\\vamsi.dasari\\Desktop\\test1\\RMA_latestrecords.xlsx',sheet_name= 'latest')

                sponsor_data1 = sponsor1.groupby(by = sponsor1['Sponsor'])
                print("....")
                for group_n,group_d in sponsor_data1:
                    #print(group_name)
                    #print(group_data)
                    study = group_d.groupby(by = group_d["Study"])
                    for group_name,group_data in study:
                        path = 'C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy'
                        entries = os.listdir(path)
                        #print(entries)
                        stri = group_name
                        name = stri.split('/')
                        group = (str(group_n)+ str(name[0]) +'.xlsx')
                        #print(group)
                        if(group in entries):
                            check = pd.ExcelFile('C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy\\'+str(group))
                            sheetnames = check.sheet_names
                            if('RMA Termination Return Orders' in sheetnames):
                                writer = pd.ExcelWriter(("C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy\\"+ str(group)),engine = 'openpyxl')
                                writer.book = load_workbook("C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy\\"+ str(group))
                                writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
                                reader = pd.read_excel(("C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy\\"+ str(group)),sheet_name = 'RMA Termination Return Orders')
                                read = reader.index
                                group_data.to_excel(writer,index=False,header=False,sheet_name = 'RMA Termination Return Orders',startrow=len(read)+1)
                                writer.close()
                        else:
                            src = 'C:\\Users\\vamsi.dasari\\Desktop\\test1\\summary_basesheet.xlsx'
                            dst = ('C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy\\' + str(group))
                            shutil.copyfile(src, dst)
                            with pd.ExcelWriter(('C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy\\'+ str(group)),engine='openpyxl', mode='a') as writer:
                                group_data.to_excel(writer, sheet_name='RMA Termination Return Orders',index=False)
                print("Created Sponser report with RMA Termination Tab")
                print("######################################")
                
                print("Creating Sponser report with Shipped Sent Items Tab")
                print("....")
                DF=pd.read_excel(shipped_items ,sheet_name="Orders Shipped")

                shipped= DF[["Shipment Date","Order #","Order Sub-Category","Study","Shipment Waybill #","Site Name","Site Number","Product Name","Sponsor"]]

                shipped['Shipment Date']=pd.to_datetime(shipped['Shipment Date'])

                SHIP = shipped.sort_values(by = "Shipment Date")

                group1 = SHIP.groupby(by = SHIP['Shipment Date'].dt.date)

                for dat,sponsor in group1:
                    #print(dat)
                    if( str(dat) == str(yesterday)):
                        #print(sponsor)
                        latest = pd.DataFrame(sponsor)
    
                #print(latest)
                try:
                    with pd.ExcelWriter(('C:\\Users\\vamsi.dasari\\Desktop\\test1\\sent_latestrecords.xlsx' )) as writer:
                        latest.to_excel(writer, sheet_name='sheet1',index=False)
                    rec1 = pd.read_excel('C:\\Users\\vamsi.dasari\\Desktop\\test1\\sent_latestrecords.xlsx',sheet_name= 'sheet1')
                except:
                    print("NO SHIPPING SENT ITEMS")
                else:
                    index_names1 = rec1[ rec1['Product Name'] == 'IQVIA Simon TMO SIM Card' ].index
                    rec1.drop(index_names1, inplace = True)
                    with pd.ExcelWriter(('C:\\Users\\vamsi.dasari\\Desktop\\test1\\sent_latestrecords.xlsx' )) as writer:
                        rec1.to_excel(writer, sheet_name='latest',index=False)
                    sponsor_ship = pd.read_excel('C:\\Users\\vamsi.dasari\\Desktop\\test1\\sent_latestrecords.xlsx',sheet_name= 'latest')

                    sponsor_data_ship = sponsor_ship.groupby(by = sponsor_ship['Sponsor'])

                    for group_n,group_d in sponsor_data_ship:
                        #print(group_name)
                        #print(group_data)
                        study = group_d.groupby(by = group_d["Study"])
                        for group_name,group_data in study:
                            path = 'C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy'
                            entries = os.listdir(path)
                            #print(entries)
                            stri = group_name
                            name = stri.split('/')
                            group = (str(group_n)+ str(name[0]) +'.xlsx')
                            #print(group)
                            if(group in entries):
                                check = pd.ExcelFile('C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy\\'+str(group))
                                sheetnames = check.sheet_names
                                if('Completed Send Items Orders' in sheetnames):
                                    writer = pd.ExcelWriter(("C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy\\"+ str(group)),engine = 'openpyxl')
                                    writer.book = load_workbook("C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy\\"+ str(group))
                                    writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
                                    reader = pd.read_excel(("C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy\\"+ str(group)),sheet_name = 'Completed Send Items Orders')
                                    read = reader.index
                                    group_data.to_excel(writer,index=False,header=False,sheet_name = 'Completed Send Items Orders',startrow=len(read)+1)
                                    writer.close()
                            else:
                                src = 'C:\\Users\\vamsi.dasari\\Desktop\\test1\\summary_basesheet.xlsx'
                                dst = ('C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy\\' + str(group))
                                shutil.copyfile(src, dst)
                                with pd.ExcelWriter(('C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy\\'+ str(group)),engine='openpyxl', mode='a') as writer:
                                    group_data.to_excel(writer, sheet_name='Completed Send Items Orders',index=False)
                    print("....")
                    print("Created Sponser report with Shipped Sent Items Tab")
        else:
            index_names = rec[ rec['Product Name'] == 'IQVIA Simon TMO SIM Card' ].index
            rec.drop(index_names, inplace = True)
            with pd.ExcelWriter(('C:\\Users\\vamsi.dasari\\Desktop\\test1\\latestrecords.xlsx' )) as writer:
                rec.to_excel(writer, sheet_name='latest',index=False)
        
            sponsor = pd.read_excel('C:\\Users\\vamsi.dasari\\Desktop\\test1\\latestrecords.xlsx',sheet_name= 'latest')

            sponsor_data = sponsor.groupby(by = sponsor['Sponsor'])
            print("....")
            file_list=[]
            for group_n,group_d in sponsor_data:
                #print(group_n)
                #print(group_d)
                study = group_d.groupby(by = group_d["Study"])
                for group_name,group_data in study:
                    path = 'C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy'
                    entries = os.listdir(path)
                    #print(entries)
                    stri = group_name
                    name = stri.split('/')
                    group = (str(group_n)+ str(name[0])+'.xlsx')
                    file_list.append(group)
                    #print(group)
                    if(group in entries):
                        writer = pd.ExcelWriter(("C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy\\"+ str(group)),engine = 'openpyxl')
                        writer.book = load_workbook("C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy\\"+ str(group))
                        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
                        reader = pd.read_excel(("C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy\\"+ str(group)),sheet_name = 'completed orders Tracking')
                        read = reader.index
                        group_data.to_excel(writer,index=False,header=False,sheet_name = 'completed orders Tracking',startrow=len(read)+1)
                        writer.close()
                    else:
                        src = 'C:\\Users\\vamsi.dasari\\Desktop\\test1\\summary_basesheet.xlsx'
                        dst = ('C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy\\' + str(group))
                        shutil.copyfile(src, dst)
                        with pd.ExcelWriter(('C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy\\'+ str(group)),engine='openpyxl', mode='a') as writer:
                            group_data.to_excel(writer, sheet_name='completed orders Tracking',index=False)
            print(file_list)
            print("....")
            print("Sponser report is created with Completed Orders Tracking Tab")
            print("######################################")
            ###########################Summary Tab Updation ###############################################
            print("Updating Sponser report with Summary Tab")
            print("....")
            for group_n,group_d in sponsor_data:
                #print(group_n)
                #print(group_d)
                study = group_d.groupby(by = group_d["Study"])
                for group_name,group_data in study:
                    path = 'C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy'
                    entries = os.listdir(path)
                    #print(entries)
                    stri = group_name
                    name = stri.split('/')
                    group = (str(group_n)+ str(name[0]) +'.xlsx')
                    #print(group)
                    Order_Sub_Category = group_data.groupby('Order Sub-Category')
                    #print(Order_Sub_Category)
                    pit=0
                    sof=0
                    dal=0
                    for a,b in Order_Sub_Category:
                        #print(a)
                        if(a == "Pittston"):
                            pit = len(b)
                            #print(pit)
                        elif(a == "Sofia"):
                            sof = len(b)
                            #print(sof)
                        elif(a == "Dalian"):
                            dal = len(b)
                            #print(dal)
                        else:
                            print()
                    check = pd.ExcelFile('C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy\\'+str(group))
                    sheetnames = check.sheet_names
                    #print(sheetnames)
                    if("Summary" in sheetnames):
                        wb = load_workbook('C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy\\'+str(group))
                        ws = wb["Summary"]
                        dest=('C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy\\'+str(group))
                        ro =len(ws['A'])
                        #print(ro)
                        last =ro + 1
                        # Write the value in the cell defined by column number        
                        ship = ws.cell( last,1)
                        ship.value = yesterday
                        pittston = ws.cell( last,2)
                        pittston.value = pit
                        sofia = ws.cell( last,3)
                        sofia.value = sof
                        Dalian = ws.cell( last,4)
                        Dalian.value = dal
                        wb.save(dest)
                        wb = load_workbook('C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy\\'+str(group))
                        ws = wb["Summary"]
                        dest1=('C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy\\'+str(group))
                        new = pd.read_excel(('C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy\\'+str(group)),sheet_name='Summary')
                        #print(new)
                        pati=new['Pittston'].sum()
                        #print(pati)
                        sofi=new['Sofia'].sum()
                        #print(sofi)
                        dali=new['Dalian'].sum()
                        #print(dali)
                        s_pit = ws.cell(2,6)
                        s_pit.value = pati
                        s_sof = ws.cell(3,6)
                        s_sof.value = sofi
                        s_dal = ws.cell(4,6)
                        s_dal.value = dali
                        wb.save(dest1)
            print("....")
            print("Updated Sponser report with Summary Tab")
            print("######################################")
            ########################## Inventory Tab Updation###########################################
            print("Creating Sponser report with Inventory Tab")
            print("....")
            for group_n,group_d in sponsor_data:
                #print(group_n)
                #print(group_d)
                study = group_d.groupby(by = group_d["Study"])
                for group_name,group_data in study:
                    path = 'C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy'
                    entries = os.listdir(path)
                    #print(entries)
                    stri = group_name
                    name = stri.split('/')
                    group = (str(group_n)+ str(name[0]) +'.xlsx')
                    #print(group)
                    check = pd.ExcelFile('C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy\\'+str(group))
                    sheetnames = check.sheet_names
                    #print(sheetnames)
                    if("Inventory" in sheetnames):
                        df1=pd.read_excel('C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy\\'+str(group),sheet_name='completed orders Tracking')
                        product = df1[['Product Name']]            
                        product_name =  product.groupby(by = product['Product Name'])
                        #print(product_name)
                        for i,j in product_name:
                            #print(i)
                            #print(j)
                            df2=pd.read_excel(inventoty_file ,sheet_name="Pivot")
                            # load excel with its path
                            wb = load_workbook(inventoty_file)
                            # to get the active work sheet
                            sh = wb.active
                            # to print the maximum number of occupied rows in console
                            max_row = sh.max_row

                            max_column = 4
                            ro=3
                            co=1
                            for r in range(1,max_row+1):
                                # to check the value in column 1
                                if(sh.cell(row=r, column=1).value) == i:
                                    for c in range(1,max_column+1):
                                        output = str(sh.cell(row=r, column=c).value)
                                        #print(output)
                                        wb = load_workbook('C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy\\'+str(group))
                                        ws = wb["Inventory"]
                                        dest1=('C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy\\'+str(group))
                                        ship = ws.cell( ro,co)
                                        ship.value = output
                                        wb.save(dest1)
                                        co=co+1
                            ro=ro+1
                            DF=pd.read_excel(inventoty_file ,sheet_name="Pivot")
                            roo=0
                            #print(DF)
                            list =[]

                            for row in DF.index:
                                if(DF['Devices (Shell Line Items)'][row] =="Accessories " ):
                                    roo = row
                                    roo=(roo+2)
                                    #print(roo)
                            for roww in DF.index:
                                dup_list=[]
                                dup_list.clear()
                                if(roww >= roo):
                                    dup_list.append(DF['Devices (Shell Line Items)'][roww])
                                    dup_list.append(DF['Unnamed: 1'][roww])
                                    dup_list.append(DF['Unnamed: 2'][roww])
                                    dup_list.append(DF['Unnamed: 3'][roww])
                                    list.append(dup_list)
                            new_data = pd.DataFrame(list,columns = ['PRODUCT', 'PHS Warehouse', 'SOC Warehouse', 'DOC Warehouse'])
                            if(group in entries):
                                writer = pd.ExcelWriter(("C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy\\"+ str(group)),engine = 'openpyxl')
                                writer.book = load_workbook("C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy\\"+ str(group))
                                writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
                                reader = pd.read_excel(("C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy\\"+ str(group)),sheet_name = 'Inventory')
                                read = reader.index
                                new_data.to_excel(writer,index=False,header=False,sheet_name = 'Inventory',startrow=8)
                                writer.close()
            print("....")
            print("Created Sponser report with Inventory Tab")
            print("######################################")
            #####################################RMA Termination Tab Updation#################################
            print("Creating Sponser report with RMA Termination Tab")
            print("....")
            DF=pd.read_excel(termination_file ,sheet_name="Order Contact")

            Rma_ship= DF[["Line Item Ship Date","Order #","Order Sub-Category","Sponsor","Study","Site Number","Product Name"]]

            Rma_ship['Line Item Ship Date']=pd.to_datetime(Rma_ship['Line Item Ship Date'])

            RMA = Rma_ship.sort_values(by = "Line Item Ship Date")

            group = RMA.groupby(by = RMA['Line Item Ship Date'].dt.date)
            print("....")
            for dat,sponsor in group:
                #print(dat)
                if( str(dat) == str(yesterday)):
                    #print(sponsor)
                    latest5 = pd.DataFrame(sponsor)
    
            print(latest5)
            try:
            
                with pd.ExcelWriter(('C:\\Users\\vamsi.dasari\\Desktop\\test1\\RMA_latestrecords.xlsx' )) as writer:
                    latest5.to_excel(writer, sheet_name='sheet1',index=False)
                rec3 = pd.read_excel('C:\\Users\\vamsi.dasari\\Desktop\\test1\\RMA_latestrecords.xlsx',sheet_name= 'sheet1')
            except:
                print("NO RMA TERMINATION RETURN ORDERS")
                try:
                    print("Creating Sponser report with Shipped Sent Items Tab")
                    print("....")
                    DF=pd.read_excel(shipped_items ,sheet_name="Orders Shipped")

                    shipped= DF[["Shipment Date","Order #","Order Sub-Category","Study","Shipment Waybill #","Site Name","Site Number","Product Name","Sponsor"]]

                    shipped['Shipment Date']=pd.to_datetime(shipped['Shipment Date'])

                    SHIP = shipped.sort_values(by = "Shipment Date")

                    group1 = SHIP.groupby(by = SHIP['Shipment Date'].dt.date)

                    for dat,sponsor in group1:
                        #print(dat)
                        if( str(dat) == str(yesterday)):
                            #print(sponsor)
                            latest6 = pd.DataFrame(sponsor)
    
                    #print(latest)
                    with pd.ExcelWriter(('C:\\Users\\vamsi.dasari\\Desktop\\test1\\sent_latestrecords.xlsx' )) as writer:
                        latest6.to_excel(writer, sheet_name='sheet1',index=False)
                    rec4 = pd.read_excel('C:\\Users\\vamsi.dasari\\Desktop\\test1\\sent_latestrecords.xlsx',sheet_name= 'sheet1')
                except:
                    print("NO SHIPPING SENT ITEMS")
                else:
                    index_names1 = rec4[ rec4['Product Name'] == 'IQVIA Simon TMO SIM Card' ].index
                    rec4.drop(index_names1, inplace = True)
                    with pd.ExcelWriter(('C:\\Users\\vamsi.dasari\\Desktop\\test1\\sent_latestrecords.xlsx' )) as writer:
                        rec.to_excel(writer, sheet_name='latest',index=False)
                    sponsor_ship = pd.read_excel('C:\\Users\\vamsi.dasari\\Desktop\\test1\\sent_latestrecords.xlsx',sheet_name= 'latest')

                    sponsor_data_ship = sponsor_ship.groupby(by = sponsor_ship['Sponsor'])

                    for group_n,group_d in sponsor_data_ship:
                        #print(group_name)
                        #print(group_data)
                        study = group_d.groupby(by = group_d["Study"])
                        for group_name,group_data in study:
                            path = 'C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy'
                            entries = os.listdir(path)
                            #print(entries)
                            stri = group_name
                            name = stri.split('/')
                            group = (str(group_n)+ str(name[0]) +'.xlsx')
                            #print(group)
                            if(group in entries):
                                check = pd.ExcelFile('C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy\\'+str(group))
                                sheetnames = check.sheet_names
                                if('Completed Send Items Orders' in sheetnames):
                                    writer = pd.ExcelWriter(("C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy\\"+ str(group)),engine = 'openpyxl')
                                    writer.book = load_workbook("C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy\\"+ str(group))
                                    writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
                                    reader = pd.read_excel(("C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy\\"+ str(group)),sheet_name = 'Completed Send Items Orders')
                                    read = reader.index
                                    group_data.to_excel(writer,index=False,header=False,sheet_name = 'Completed Send Items Orders',startrow=len(read)+1)
                                    writer.close()
                    print("....")
                    print("Created Sponser report with Shipped Sent Items Tab")
            else:
                index_names = rec3[ rec3['Product Name'] == 'IQVIA Simon TMO SIM Card' ].index
                rec3.drop(index_names, inplace = True)
                with pd.ExcelWriter(('C:\\Users\\vamsi.dasari\\Desktop\\test1\\RMA_latestrecords.xlsx' )) as writer:
                    rec3.to_excel(writer, sheet_name='latest',index=False)
                sponsor1 = pd.read_excel('C:\\Users\\vamsi.dasari\\Desktop\\test1\\RMA_latestrecords.xlsx',sheet_name= 'latest')

                sponsor_data1 = sponsor1.groupby(by = sponsor1['Sponsor'])
                print("....")
                for group_n,group_d in sponsor_data1:
                    #print(group_name)
                    #print(group_data)
                    study = group_d.groupby(by = group_d["Study"])
                    for group_name,group_data in study:
                        path = 'C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy'
                        entries = os.listdir(path)
                        #print(entries)
                        stri = group_name
                        name = stri.split('/')
                        group = (str(group_n)+ str(name[0]) +'.xlsx')
                        #print(group)
                        if(group in entries):
                            check = pd.ExcelFile('C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy\\'+str(group))
                            sheetnames = check.sheet_names
                            if('Completed Send Items Orders' in sheetnames):
                                writer = pd.ExcelWriter(("C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy\\"+ str(group)),engine = 'openpyxl')
                                writer.book = load_workbook("C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy\\"+ str(group))
                                writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
                                reader = pd.read_excel(("C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy\\"+ str(group)),sheet_name = 'RMA Termination Return Orders')
                                read = reader.index
                                group_data.to_excel(writer,index=False,header=False,sheet_name = 'RMA Termination Return Orders',startrow=len(read)+1)
                                writer.close()
                print("Created Sponser report with RMA Termination Tab")
                print("######################################")
                ########################## Shipped Sent Items Updation #####################################
                print("Creating Sponser report with Shipped Sent Items Tab")
                print("....")
                DF=pd.read_excel(shipped_items ,sheet_name="Orders Shipped")

                shipped= DF[["Shipment Date","Order #","Order Sub-Category","Study","Shipment Waybill #","Site Name","Site Number","Product Name","Sponsor"]]

                shipped['Shipment Date']=pd.to_datetime(shipped['Shipment Date'])

                SHIP = shipped.sort_values(by = "Shipment Date")

                group1 = SHIP.groupby(by = SHIP['Shipment Date'].dt.date)

                for dat,sponsor in group1:
                    #print(dat)
                    if( str(dat) == str(yesterday)):
                        #print(sponsor)
                        latest = pd.DataFrame(sponsor)
    
                #print(latest)
                try:
                    with pd.ExcelWriter(('C:\\Users\\vamsi.dasari\\Desktop\\test1\\sent_latestrecords.xlsx' )) as writer:
                        latest.to_excel(writer, sheet_name='sheet1',index=False)
                    rec1 = pd.read_excel('C:\\Users\\vamsi.dasari\\Desktop\\test1\\sent_latestrecords.xlsx',sheet_name= 'sheet1')
                except:
                    print("NO SHIPPED SENT ITEMS")
                else:
                    index_names1 = rec1[ rec1['Product Name'] == 'IQVIA Simon TMO SIM Card' ].index
                    rec1.drop(index_names1, inplace = True)
                    with pd.ExcelWriter(('C:\\Users\\vamsi.dasari\\Desktop\\test1\\sent_latestrecords.xlsx' )) as writer:
                        rec.to_excel(writer, sheet_name='latest',index=False)
                    sponsor_ship = pd.read_excel('C:\\Users\\vamsi.dasari\\Desktop\\test1\\sent_latestrecords.xlsx',sheet_name= 'latest')

                    sponsor_data_ship = sponsor_ship.groupby(by = sponsor_ship['Sponsor'])

                    for group_n,group_d in sponsor_data_ship:
                        #print(group_name)
                        #print(group_data)
                        study = group_d.groupby(by = group_d["Study"])
                        for group_name,group_data in study:
                            path = 'C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy'
                            entries = os.listdir(path)
                            #print(entries)
                            stri = group_name
                            name = stri.split('/')
                            group = (str(group_n)+ str(name[0]) +'.xlsx')
                            #print(group)
                            if(group in entries):
                                check = pd.ExcelFile('C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy\\'+str(group))
                                sheetnames = check.sheet_names
                                if('Completed Send Items Orders' in sheetnames):
                                    writer = pd.ExcelWriter(("C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy\\"+ str(group)),engine = 'openpyxl')
                                    writer.book = load_workbook("C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy\\"+ str(group))
                                    writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
                                    reader = pd.read_excel(("C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy\\"+ str(group)),sheet_name = 'Completed Send Items Orders')
                                    read = reader.index
                                    group_data.to_excel(writer,index=False,header=False,sheet_name = 'Completed Send Items Orders',startrow=len(read)+1)
                                    writer.close()
                print("....")
                print("Created Sponser report with Shipped Sent Items Tab")
                
                for file1 in file_list:
                    print(file1)
                    src1 = ('C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daliy\\'+str(file1))
                    dst1 = ('C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daily_mail\\'+str(file1))
                    shutil.copyfile(src1, dst1)

                os.startfile('outlook')
                time.sleep(1.0)
                path = 'C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daily_mail'
                entries = os.listdir(path)
                for filename in entries:
                    Att_path = "C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daily_mail\\"+filename
                    name = filename.split('.')
                    firstname = name[0]
                    subject="IQVIA "+ str(firstname) + " Daily Report"
                    body= '''Hello All,

Attached is today’s daily report for '''+str(firstname)+'''.

Best Regards,
Emil Anev
Supervisor
Mobile: +359876955560; Office: +(1) 973 585 1841
Email Address: emil.anev@hcl.com
C3i Solutions, an HCL Technologies Company
Business Park Sofia,Bld. 7, fl.1
Sofia, Bulgaria, 1766 | www.c3isolutions.com'''

                    outlook=client.Dispatch("Outlook.Application")
                    mail_from=outlook.session.Accounts['vamsi.dasari@hcl.com']

                    message=outlook.createItem(0)
                    message.Display()
                    message.To="devilalithamba.paidi@hcl.com"
                    #message.CC="basireddygari.suguna@hcl.com"
                    message.Subject= subject
                    message.Attachments.Add(Att_path)
                    message.Body= body
                    #message._oleobj_.Invoke(*(64209, 0, 8,gmail))
                    ##message.Save()
                    message.Send()

                time.sleep(1.0)
                MsgBox1 = messagebox.askquestion('Exit Application','Successfully sent mail,please close the widget',icon = 'warning')
                shutil.rmtree('C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daily_mail')
                os.mkdir('C:\\Users\\vamsi.dasari\\Desktop\\IQVIA_Daily_mail')
    
   
text = Label(text="Once files are selected, Click On Generate Report Button",font = "Helvetica 12 bold italic")
text.place(x=75,y=330)

GenReportButton = Button(text="Generate Report",fg="Green",font = "Helvetica 12 bold italic",command=GenerateReport)
GenReportButton.place(x = 130,y = 380)    

def close():         
    raise SystemExit

ExitButton = Button(text="  Exit  ",fg="Red",font = "Helvetica 12 bold italic",command=close)
ExitButton.place(x = 340,y = 380)

root.mainloop()
